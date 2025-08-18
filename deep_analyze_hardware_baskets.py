#!/usr/bin/env python3
import pandas as pd
import sys
import openpyxl
import json
from collections import defaultdict

def deep_analyze_excel_file(filepath):
    """
    Comprehensive analysis of hardware basket Excel files
    """
    print(f"\n{'='*80}")
    print(f"🔍 DEEP ANALYSIS: {filepath}")
    print(f"{'='*80}")
    
    try:
        # Load the workbook to see all sheet names
        wb = openpyxl.load_workbook(filepath, read_only=True)
        print(f"📋 Total sheets: {len(wb.sheetnames)}")
        print(f"📋 Sheet names: {wb.sheetnames}")
        
        file_structure = {
            "file_path": filepath,
            "total_sheets": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "sheets": {}
        }
        
        # Analyze each sheet in detail
        for sheet_name in wb.sheetnames:
            print(f"\n🔍 ANALYZING SHEET: '{sheet_name}'")
            print(f"-" * 60)
            
            try:
                # Read the sheet - no row limit to see full structure
                df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
                
                sheet_info = {
                    "dimensions": f"{df.shape[0]} rows x {df.shape[1]} columns",
                    "total_rows": df.shape[0],
                    "total_cols": df.shape[1],
                    "non_empty_rows": 0,
                    "header_candidates": [],
                    "data_patterns": {},
                    "sample_data": {}
                }
                
                print(f"  📏 Dimensions: {df.shape[0]} rows x {df.shape[1]} columns")
                
                # Find non-empty rows
                non_empty_mask = ~df.isnull().all(axis=1)
                non_empty_rows = df[non_empty_mask].index.tolist()
                sheet_info["non_empty_rows"] = len(non_empty_rows)
                print(f"  📊 Non-empty rows: {len(non_empty_rows)}")
                
                # Look for potential headers by analyzing string patterns
                potential_headers = []
                for i in range(min(20, df.shape[0])):  # Check first 20 rows for headers
                    row = df.iloc[i]
                    non_null_values = row.dropna()
                    if len(non_null_values) > 3:  # At least 4 columns with data
                        string_count = sum(1 for val in non_null_values if isinstance(val, str) and len(str(val)) > 2)
                        if string_count >= len(non_null_values) * 0.7:  # 70% strings
                            potential_headers.append({
                                "row": i,
                                "values": non_null_values.tolist()[:10],  # First 10 values
                                "string_ratio": string_count / len(non_null_values)
                            })
                
                sheet_info["header_candidates"] = potential_headers
                print(f"  🏷️  Potential header rows: {len(potential_headers)}")
                for header in potential_headers[:3]:  # Show first 3
                    print(f"    Row {header['row']}: {header['values'][:5]}...")
                
                # Analyze data patterns in different sections
                data_sections = []
                current_section_start = None
                
                for i in range(df.shape[0]):
                    row = df.iloc[i]
                    non_null_count = row.count()
                    
                    if non_null_count > 3:  # Significant data
                        if current_section_start is None:
                            current_section_start = i
                    else:
                        if current_section_start is not None:
                            # End of section
                            data_sections.append((current_section_start, i-1))
                            current_section_start = None
                
                # Close last section if needed
                if current_section_start is not None:
                    data_sections.append((current_section_start, df.shape[0]-1))
                
                print(f"  📊 Data sections found: {len(data_sections)}")
                for j, (start, end) in enumerate(data_sections[:5]):  # Show first 5 sections
                    section_rows = end - start + 1
                    print(f"    Section {j+1}: Rows {start}-{end} ({section_rows} rows)")
                    
                    # Sample data from this section
                    sample_row = df.iloc[start:start+3].fillna('')  # First few rows of section
                    sheet_info["sample_data"][f"section_{j+1}"] = sample_row.to_dict('records')
                
                # Look for pricing/numeric patterns
                numeric_columns = []
                for col in df.columns:
                    numeric_count = pd.to_numeric(df[col], errors='coerce').count()
                    if numeric_count > 10:  # At least 10 numeric values
                        numeric_columns.append({
                            "column": col,
                            "numeric_count": numeric_count,
                            "sample_values": pd.to_numeric(df[col], errors='coerce').dropna().head(5).tolist()
                        })
                
                sheet_info["numeric_columns"] = numeric_columns
                print(f"  💰 Columns with significant numeric data: {len(numeric_columns)}")
                
                # Look for vendor/product indicators
                vendor_indicators = ['dell', 'lenovo', 'hpe', 'server', 'rack', 'intel', 'amd', 'processor', 'cpu', 'memory', 'storage']
                content_analysis = {}
                
                for indicator in vendor_indicators:
                    count = 0
                    for col in df.columns:
                        count += df[col].astype(str).str.lower().str.contains(indicator, na=False).sum()
                    if count > 0:
                        content_analysis[indicator] = count
                
                sheet_info["content_indicators"] = content_analysis
                if content_analysis:
                    print(f"  🏷️  Content indicators: {content_analysis}")
                
                # Store detailed sample of first 10 rows for manual inspection
                sheet_info["first_10_rows"] = df.head(10).fillna('').to_dict('records')
                
                file_structure["sheets"][sheet_name] = sheet_info
                
            except Exception as e:
                print(f"  ❌ Error analyzing sheet '{sheet_name}': {e}")
                file_structure["sheets"][sheet_name] = {"error": str(e)}
        
        wb.close()
        
        # Save detailed analysis to JSON
        json_filename = filepath.replace('.xlsx', '_analysis.json').replace(' ', '_')
        with open(json_filename, 'w') as f:
            json.dump(file_structure, f, indent=2, default=str)
        
        print(f"\n💾 Detailed analysis saved to: {json_filename}")
        
        return file_structure
        
    except Exception as e:
        print(f"❌ Error analyzing file: {e}")
        return None

def compare_file_structures(dell_analysis, lenovo_analysis):
    """
    Compare the structures of Dell and Lenovo files to find patterns
    """
    print(f"\n{'='*80}")
    print(f"🔄 COMPARATIVE ANALYSIS: Dell vs Lenovo")
    print(f"{'='*80}")
    
    if not dell_analysis or not lenovo_analysis:
        print("❌ Cannot compare - one or both analyses failed")
        return
    
    print(f"📊 Dell file has {dell_analysis['total_sheets']} sheets")
    print(f"📊 Lenovo file has {lenovo_analysis['total_sheets']} sheets")
    
    # Compare sheet names
    dell_sheets = set(dell_analysis['sheet_names'])
    lenovo_sheets = set(lenovo_analysis['sheet_names'])
    
    common_sheets = dell_sheets & lenovo_sheets
    dell_only = dell_sheets - lenovo_sheets
    lenovo_only = lenovo_sheets - dell_sheets
    
    print(f"\n📋 Common sheets: {list(common_sheets)}")
    print(f"📋 Dell-only sheets: {list(dell_only)}")
    print(f"📋 Lenovo-only sheets: {list(lenovo_only)}")
    
    # Analyze common sheets for structural similarities
    for sheet_name in common_sheets:
        if sheet_name in dell_analysis['sheets'] and sheet_name in lenovo_analysis['sheets']:
            dell_sheet = dell_analysis['sheets'][sheet_name]
            lenovo_sheet = lenovo_analysis['sheets'][sheet_name]
            
            print(f"\n🔍 Comparing sheet: '{sheet_name}'")
            print(f"  Dell: {dell_sheet.get('dimensions', 'unknown')}")
            print(f"  Lenovo: {lenovo_sheet.get('dimensions', 'unknown')}")
            
            # Compare header candidates
            dell_headers = dell_sheet.get('header_candidates', [])
            lenovo_headers = lenovo_sheet.get('header_candidates', [])
            
            if dell_headers and lenovo_headers:
                print(f"  Header patterns:")
                print(f"    Dell potential headers at rows: {[h['row'] for h in dell_headers[:3]]}")
                print(f"    Lenovo potential headers at rows: {[h['row'] for h in lenovo_headers[:3]]}")

if __name__ == "__main__":
    # Analyze both files - using actual file paths in the workspace
    dell_file = "/Users/mateimarcu/DevApps/LCMDesigner/docs/X86 Basket Q3 2025 v2 Dell Only.xlsx"
    lenovo_file = "/Users/mateimarcu/DevApps/LCMDesigner/docs/X86 Basket Q3 2025 v2 Lenovo Only.xlsx"
    
    print("🚀 Starting comprehensive hardware basket analysis...")
    
    dell_analysis = deep_analyze_excel_file(dell_file)
    lenovo_analysis = deep_analyze_excel_file(lenovo_file)
    
    # Compare the structures
    compare_file_structures(dell_analysis, lenovo_analysis)
    
    print(f"\n✅ Analysis complete!")
    print(f"📁 Check the generated *_analysis.json files for detailed data structure information")
