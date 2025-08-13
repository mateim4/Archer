#!/usr/bin/env python3
import pandas as pd
import sys
import openpyxl
from openpyxl import load_workbook
import json

def analyze_excel_structure(filepath):
    print(f"📊 COMPREHENSIVE ANALYSIS OF: {filepath}")
    print("="*80)
    
    try:
        # Load the workbook to see all sheet names
        wb = load_workbook(filepath, read_only=True, data_only=True)
        print(f"📋 Total sheets: {len(wb.sheetnames)}")
        print(f"📋 Sheet names: {wb.sheetnames}")
        print("\n")
        
        # Analyze each sheet comprehensively
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            print(f"🔍 SHEET {sheet_idx + 1}: '{sheet_name}'")
            print("-" * 60)
            
            try:
                # Get the worksheet
                ws = wb[sheet_name]
                
                # Get dimensions
                max_row = ws.max_row
                max_col = ws.max_column
                print(f"  📏 Dimensions: {max_row} rows x {max_col} columns")
                
                # Read with pandas for better data analysis
                df = pd.read_excel(filepath, sheet_name=sheet_name, header=None, nrows=min(100, max_row))
                
                print(f"  📊 Data shape: {df.shape}")
                
                # Find non-empty rows and columns
                non_empty_rows = []
                non_empty_cols = []
                
                for row_idx in range(min(50, len(df))):
                    row_data = df.iloc[row_idx]
                    if not row_data.isna().all():
                        non_empty_rows.append(row_idx)
                        
                for col_idx in range(len(df.columns)):
                    col_data = df.iloc[:, col_idx]
                    if not col_data.isna().all():
                        non_empty_cols.append(col_idx)
                
                print(f"  📄 Non-empty rows (first 50): {len(non_empty_rows)} rows with data")
                print(f"  📄 Non-empty columns: {len(non_empty_cols)} columns with data")
                
                # Analyze header patterns
                print(f"\n  🔍 HEADER ANALYSIS:")
                for row_idx in range(min(10, len(df))):
                    row_data = df.iloc[row_idx]
                    non_null_values = [str(val) for val in row_data if pd.notna(val) and str(val).strip()]
                    if non_null_values:
                        print(f"    Row {row_idx}: {non_null_values[:8]}{'...' if len(non_null_values) > 8 else ''}")
                
                # Look for potential data patterns
                print(f"\n  🔍 DATA PATTERN ANALYSIS:")
                
                # Find rows that look like headers (contain many text values)
                potential_headers = []
                for row_idx in range(min(20, len(df))):
                    row_data = df.iloc[row_idx]
                    text_count = sum(1 for val in row_data if pd.notna(val) and isinstance(val, str) and len(str(val)) > 2)
                    if text_count >= 3:  # Row has multiple text values
                        potential_headers.append((row_idx, text_count))
                
                if potential_headers:
                    print(f"    Potential header rows: {[f'Row {r} ({t} text cols)' for r, t in potential_headers[:5]]}")
                
                # Find rows with numeric data (potential data rows)
                numeric_rows = []
                for row_idx in range(min(50, len(df))):
                    row_data = df.iloc[row_idx]
                    numeric_count = sum(1 for val in row_data if pd.notna(val) and isinstance(val, (int, float)) and val != 0)
                    if numeric_count >= 2:  # Row has multiple numeric values
                        numeric_rows.append((row_idx, numeric_count))
                
                if numeric_rows:
                    print(f"    Rows with numeric data: {len(numeric_rows)} rows")
                    print(f"    Sample numeric rows: {[f'Row {r} ({n} nums)' for r, n in numeric_rows[:5]]}")
                
                # Sample actual data from key rows
                print(f"\n  📄 SAMPLE DATA:")
                sample_rows = [0, 1, 2, 3, 5, 10, 15, 20] if len(df) > 20 else list(range(min(10, len(df))))
                
                for row_idx in sample_rows:
                    if row_idx < len(df):
                        row_data = df.iloc[row_idx]
                        # Show first 6 meaningful values
                        meaningful_values = []
                        for col_idx, val in enumerate(row_data):
                            if pd.notna(val) and str(val).strip():
                                meaningful_values.append(f"Col{col_idx}='{val}'")
                            if len(meaningful_values) >= 6:
                                break
                        
                        if meaningful_values:
                            print(f"    Row {row_idx}: {', '.join(meaningful_values)}")
                
                # Look for pricing patterns
                print(f"\n  💰 PRICING ANALYSIS:")
                price_patterns = []
                for row_idx in range(min(30, len(df))):
                    row_data = df.iloc[row_idx]
                    for col_idx, val in enumerate(row_data):
                        if pd.notna(val) and isinstance(val, (int, float)) and val > 100:  # Potential price
                            price_patterns.append(f"Row {row_idx}, Col {col_idx}: {val}")
                
                if price_patterns:
                    print(f"    Potential prices found: {len(price_patterns)} values")
                    print(f"    Sample prices: {price_patterns[:5]}")
                
                # Look for model/product patterns
                print(f"\n  🖥️  PRODUCT/MODEL ANALYSIS:")
                product_patterns = []
                for row_idx in range(min(30, len(df))):
                    row_data = df.iloc[row_idx]
                    for col_idx, val in enumerate(row_data):
                        if pd.notna(val) and isinstance(val, str):
                            val_str = str(val).strip()
                            # Look for server/hardware terms
                            if any(term in val_str.lower() for term in ['server', 'rack', 'proc', 'intel', 'amd', 'dell', 'r450', 'r650']):
                                product_patterns.append(f"Row {row_idx}, Col {col_idx}: '{val_str}'")
                
                if product_patterns:
                    print(f"    Product mentions found: {len(product_patterns)} entries")
                    print(f"    Sample products: {product_patterns[:5]}")
                
            except Exception as e:
                print(f"  ❌ Error analyzing sheet '{sheet_name}': {e}")
            
            print("\n" + "="*80 + "\n")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error analyzing file: {e}")

def analyze_specific_sheet_detailed(filepath, sheet_name):
    print(f"\n🔍 DETAILED ANALYSIS OF SHEET: '{sheet_name}'")
    print("="*80)
    
    try:
        # Read the full sheet
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        print(f"📊 Full sheet dimensions: {df.shape}")
        
        # Analyze column by column for the first 20 rows
        print(f"\n📋 COLUMN-BY-COLUMN ANALYSIS (First 20 rows):")
        for col_idx in range(min(20, df.shape[1])):
            col_data = df.iloc[:20, col_idx]
            non_null_values = [str(val) for val in col_data if pd.notna(val) and str(val).strip()]
            
            if non_null_values:
                print(f"  Column {col_idx}: {len(non_null_values)} values")
                print(f"    Sample: {non_null_values[:5]}")
                
                # Detect data types
                numeric_count = sum(1 for val in non_null_values if val.replace('.', '').replace(',', '').replace('-', '').isdigit())
                text_count = len(non_null_values) - numeric_count
                print(f"    Types: {numeric_count} numeric, {text_count} text")
                print()
        
    except Exception as e:
        print(f"❌ Error in detailed analysis: {e}")

if __name__ == "__main__":
    test_file = "legacy-server/test-basket.xlsx"
    
    print("🔍 HARDWARE BASKET FILE ANALYSIS")
    print("="*80)
    
    analyze_excel_structure(test_file)
    
    # Do detailed analysis of the main data sheet
    analyze_specific_sheet_detailed(test_file, "Dell Lot Pricing")
