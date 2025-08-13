#!/usr/bin/env python3
import pandas as pd
import sys
import openpyxl

def analyze_excel_file(filepath):
    print(f"📊 Analyzing Excel file: {filepath}")
    
    try:
        # Load the workbook to see all sheet names
        wb = openpyxl.load_workbook(filepath, read_only=True)
        print(f"📋 Sheet names: {wb.sheetnames}")
        
        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            print(f"\n🔍 Analyzing sheet: {sheet_name}")
            
            try:
                # Read the sheet with pandas
                df = pd.read_excel(filepath, sheet_name=sheet_name, nrows=20)
                
                print(f"  📏 Dimensions: {df.shape[0]} rows x {df.shape[1]} columns")
                print(f"  📊 Columns: {list(df.columns)}")
                
                # Show first few rows
                print(f"  📄 First 5 rows:")
                for i, row in df.head().iterrows():
                    print(f"    Row {i}: {dict(row)}")
                
                # Look for potential hardware data indicators
                column_text = ' '.join(str(col).lower() for col in df.columns)
                if any(keyword in column_text for keyword in ['model', 'server', 'hardware', 'cpu', 'memory', 'disk', 'price', 'part']):
                    print(f"  ✅ Sheet '{sheet_name}' appears to contain hardware data!")
                else:
                    print(f"  ❓ Sheet '{sheet_name}' may not contain hardware data")
                    
            except Exception as e:
                print(f"  ❌ Error reading sheet '{sheet_name}': {e}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error analyzing file: {e}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        analyze_excel_file(sys.argv[1])
    else:
        # Analyze available test files
        test_files = [
            "legacy-server/test-basket.xlsx",
            "test-files/test-hardware-basket.xlsx"
        ]
        
        for test_file in test_files:
            try:
                analyze_excel_file(test_file)
                print("\n" + "="*80 + "\n")
            except FileNotFoundError:
                print(f"⚠️  File not found: {test_file}")
